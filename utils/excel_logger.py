import os
import openpyxl
from openpyxl import Workbook
from datetime import datetime
import re
from typing import List, Dict, Any
from utils.logger import get_logger

logger = get_logger(__name__)

LOG_FILE = "data/veritabani_log.xlsx"

HEADERS = [
    "Yükleme Tarihi & Saati",
    "Dosya Adı",
    "Tespit Edilen Bölge",
    "Okunan Lisans No",
    "Karttaki İsim",
    "Eşleşme Durumu",
    "Excel'deki Sıra",
    "Doğrulama Mesajı / Eşleşen İsim",
    "Kaynak Linki"
]

def append_logs(results: List[Dict[str, Any]]) -> None:
    """
    Belirtilen sonuç listesini Excel log dosyasına (append mantığıyla) ekler.
    Dosya yoksa oluşturur.
    """
    if not results:
        return
        
    # Data klasörünün var olduğundan emin ol
    os.makedirs(os.path.dirname(LOG_FILE), exist_ok=True)
    
    file_exists = os.path.exists(LOG_FILE)
    
    if file_exists:
        try:
            wb = openpyxl.load_workbook(LOG_FILE)
            ws = wb.active
        except Exception as e:
            logger.error(f"Mevcut log dosyası okunamadı, yenisi oluşturuluyor: {e}")
            wb = Workbook()
            ws = wb.active
            ws.append(HEADERS)
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(HEADERS)
        
    for result in results:
        filename = result.get("filename", "")
        upload_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        extracted_data = result.get("extracted_data", {})
        
        detected_region = result.get("detected_region_name") or extracted_data.get("region", "")
        license_no = result.get("ocr_license_found") or extracted_data.get("license_no", "")
        card_name = extracted_data.get("name", "")
        
        # status and label can be mapped
        status = result.get("label") or result.get("status", "")
        message = result.get("message", "")
        source_link = result.get("url", "")
        
        # Mesajın içinden Excel sırasını yakalamaya çalış
        excel_row = ""
        row_match = re.search(r"Excel listesinde (\d+)\. sıradaki", message)
        if row_match:
            excel_row = row_match.group(1)
            
        row = [
            upload_time,
            filename,
            detected_region,
            license_no,
            card_name,
            status,
            excel_row,
            message,
            source_link
        ]
        
        # Tüm hücreleri string yapalım ki Excel kütüphanesi uyarı vermesin
        row = [str(item) if item is not None else "" for item in row]
        
        ws.append(row)
        
    try:
        wb.save(LOG_FILE)
        logger.info(f"{len(results)} kayıt başarıyla {LOG_FILE} dosyasına eklendi.")
    except Exception as e:
        logger.error(f"Log dosyası kaydedilirken hata oluştu: {e}")
